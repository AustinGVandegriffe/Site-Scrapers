import os
path = '.'
os.chdir(path)
try:
    os.mkdir('Selenium_Download_TMP')
except:
    pass

# BeautifulSoup can be replaced with XPath searches with Selenium
from bs4 import BeautifulSoup as bs

import openpyxl as xl
import sqlite3

import selenium
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains

from urllib.parse import urljoin
import time
import subprocess
import keyboard


options = webdriver.ChromeOptions()
profile = {"plugins.plugins_list": [{"enabled":False,"name":"Chrome PDF Viewer"}],
        "download": {
            "prompt_for_download": False,
            "default_directory"  : f'{path}/Selenium_Download_TMP'
    }
}
options.add_experimental_option("prefs",profile)
driver = webdriver.Chrome(chrome_options = options)

def alpha(letter):
    return ord(letter)-65

def xl_to_num(loc):
    loc = loc.split(':')
    if len(loc) == 1:
        col = loc[0]
        if len(col) == 1:
            return alpha(col)+1
        else:
            return (alpha(col[0])+1)*26+alpha(col[1])+1
    else:
        col,row = loc
        row = int(row)
        if len(col) == 1:
            return (row,alpha(col)+1)
        else:
            return (row,(alpha(col[0])+1)*26+alpha(col[1])+1)

def aem_asm(current_url,soup):
    return urljoin(current_url,soup.find_all('a')[0])

def bio_one(current_url,soup):
    tags = soup.find('div',{'id':'contentSidebarWrapper'}).find_all('a')
    tmp = [x for x in tags if 'PDF' == x.text][0]['href']
    return urljoin(current_url,tmp)

def oxford(current_url,soup):
    tmp = [x for x in soup.find('ul',{'id':'Toolbar'}).find_all('a') if 'PDF' == x.text.replace(' ','').replace('\n','')][0]['href']
    return urljoin(current_url,tmp)

def springer(current_url,soup):
    try:
        tmp = soup.find_all('a')
        tmp = [x for x in tmp if 'Download PDF' in x.text.replace('\n',' ').replace('  ',' ')][0]['href']
        return urljoin(current_url,tmp)
    except TypeError:
        print(f'ERROR: {current_url}')
        return current_url

def aps(current_url,soup):
    # Protected: prompt for manual download in webdriver:
    #   https://journals.aps.org/pre/pdf/10.1103/PhysRevE.61.557
    # Different Colors all have same HTML:
    #   Red: https://journals.aps.org/prb/
    #   Orange: https://journals.aps.org/rmp/
    #   Green: https://journals.aps.org/prl/
    tmp = soup.find('div',{'id':'article-actions'}).find_all('a')
    tmp = [x for x in tmp if x.text.replace(' ','').replace('\n','')=='PDF'][0]['href']
    return urljoin(current_url,tmp)

def iopscience(current_url,soup):
    tmp = soup.find('div',{'id':'page-content'}).find('div',{'class':'btn-multi-block mb-1'}).find('a')['href']
    return urljoin(current_url,tmp)

def aapt(current_url,soup):
    tmp = soup.find('a',{'class':'pdf'})['href']
    return urljoin(current_url,tmp)

def plos(current_url,soup):
    tmp = soup.find('a',{'id':'downloadPdf'})['href']
    return urljoin(current_url,tmp)

def hindawi(current_url,soup):
    tmp = soup.find('div',{'id':'article_list'}).find('a',{'class':'full_text_pdf'})['href']
    return urljoin(current_url,tmp)

def pubs_rsc(current_url,soup):
    tmp = soup.find('div',{'id':'DownloadOption'}).find('a')
    if 'Buy this article' in tmp.text:
        return ''
    return urljoin(current_url,tmp['href'])

def nature(current_url,soup):
    tmp = soup.find_all('a')
    for x in tmp:
        if 'DownloadPDF' in x.text.replace(' ','').replace('\n',''):
            tmp = x['href']
            break
    return urljoin(current_url,tmp)

def bmc(current_url,soup):
    tmp = soup.find('a',{'id':'articlePdf'})['href']
    return urljoin(current_url,tmp)

def mdpi(current_url,soup):
    tmp = soup.find('a',{'id':'pdf_link'})['href']
    return urljoin(current_url,tmp)

def sage(current_url,soup):
    tmp = soup.find('td',{'class':'pdfBadge'}).find('a')['href']
    return urljoin(current_url,tmp)

def euclid(current_url,soup):
    try:
        time.sleep(2)
        driver.switch_to.alert().accept()
    except selenium.common.exceptions.NoAlertPresentException:
        pass
    keyboard.press_and_release('enter')

    button = driver.find_element_by_class_name('main-content').find_element_by_class_name('publication-content')
    button = [x for x in button.find_elements_by_tag_name('a') if 'pdf' in x.text.lower()][0]
    ActionChains(driver).move_to_element(button).click(button).perform()
    # return urljoin(current_url,tmp)
    while '.pdf' not in [os.path.splitext(x)[1].lower() for x in os.listdir(f'{path}/Selenium_Download_TMP')]:
        pass
    return current_url

def siam(current_url,soup):
    tmp = soup.find('ul',{'id':'articleToolList'}).find('li',{'class':'articleToolLi showPDF'}).find('a')['href']
    return urljoin(current_url,tmp)

def scirp(current_url,soup):
    # tmp = soup.find('div',{'id':'con_one_1'}).find_all('a')
    tmp = soup.find('ul',{'class':'r_nav'}).find_all('a')
    tmp = [x for x in tmp if 'Full-TextPDF' in x.text.replace(' ','').replace('\n','')][0]['href']
    return urljoin(current_url,tmp)

def uchicago(current_url,soup):
    tmp = soup.find_all('li')
    tmp = [x for x in tmp if 'PDF'==x.text.replace(' ','').replace('\n','')][0].find('a')['href']
    return urljoin(current_url,tmp)


def iumj(current_url,soup):
    tmp = soup.find('div',{'id':'subcontentb'}).find('tbody').find('a')['href']
    return urljoin(current_url,tmp)


def sciedupress(current_url,soup):
    tmp = soup.find('div',{'id':'articleFullText'}).find('a')['href']
    return urljoin(current_url,tmp)

def geoscienceworld(driver):
    soup = bs(driver.page_source,'html.parser')
    tmp = soup.find('ul',{'id':'Toolbar'}).find('li',{'class':'toolbar-item item-pdf'}).find('a')['href']
    return urljoin(driver.current_url,tmp)


def jnsbm(driver):
    soup = bs(driver.page_source,'html.parser')
    tmp = soup.find('table',{'id':'table10'}).find('a')['href']
    driver.get(urljoin(driver.current_url,tmp))
    soup = bs(driver.page_source,'html.parser')
    tmp = soup.find('table',{'id':'tsw2'}).find('a')['href']
    return urljoin(driver.current_url,tmp)

def jci(driver):
    soup = bs(driver.page_source,'html.parser')
    tmp = soup.find('ul',{'id':'article-tools'}).find_all('a')
    tmp = [x for x in tmp if 'ViewPDF' == x.text.replace(' ','').replace('\n','')][0]['href']
    driver.get(urljoin(driver.current_url,tmp))
    soup = bs(driver.page_source,'html.parser')
    tmp = [x for x in soup.find_all('h3') if 'FullTextPDF' in x.text.replace(' ','').replace('\n','')][0]
    tmp = tmp.find('a')['href']
    return urljoin(driver.current_url,tmp)

def sciencedirect(driver):
    # actions = selenium.webdriver.common.action_chains.ActionChains(driver)
    # actions.double_click()
    # actions.perform()
    # actions.move_by_offset(0,0).context_click().perform()
    try:
        driver.find_element_by_id('pdfLink').click()
        soup = bs(driver.page_source, 'html.parser')
        tmp = soup.find('div',{'id':'popover-content-download-pdf-popover'}).find('div',{'class':'PdfDropDownMenu'}).find('a')['href']
        return urljoin(driver.current_url,tmp)
    except:
        return driver.current_url

def scripts_iucr(driver):
    soup = bs(driver.page_source, 'html.parser')
    tmp = soup.find('div',{'id':'main'}).find_all('a')
    tmp = [x for x in tmp if 'Readarticle' == x.text.replace(' ','').replace('\n','').replace('\xa0','')][0]['href']
    try:
        subprocess.Popen(['python','./SUPPORT/esc.py'])
        driver.get(urljoin(driver.current_url,tmp))
        time.sleep(3)
        soup = bs(driver.page_source, 'html.parser')
        tmp_ = soup.find('div',{'id':'art_leftbox'}).find_all('a')[2]['href']
        tmp = tmp_
    except:
        pass
    return urljoin(driver.current_url,tmp)

def osa(current_url,soup):
    tmp = soup.find('div',{'id':'articleContainer'}).find('li',{'class':'pdf-download'}).find('a')['href']
    return urljoin(current_url,tmp)

def wiley(driver):
    soup = bs(driver.page_source,'html.parser')
    tmp = soup.find('a',{'title':'Article PDF'})['href']
    try:
        driver.get(urljoin(driver.current_url,tmp))
    except:
        print(f"ERROR: {driver.current_url}")

def jwsr(driver):
    soup = bs(driver.page_source, 'html.parser')
    tmp = soup.find('div',{'id':'articleFullText'}).find('a')['href']
    driver.get(urljoin(driver.current_url,tmp))
    soup = bs(driver.page_source, 'html.parser')
    tmp = soup.find('div',{'id':'pdfDownloadLinkContainer'}).find('a')['href']
    return urljoin(driver.current_url,tmp)

def main():
    global driver

    doi_col = xl_to_num('B')
    url_col = xl_to_num('C')
    issn_col = xl_to_num('A')

    conn = sqlite3.connect('./SUPPORT/journal.db')
    c = conn.cursor()
    # conn = sqlite3.connect(':memory:')

    # row_start = int(input('Start at row: '))
    row_start = 1

    for file in os.listdir("./PROCESS"):
        name, ext = os.path.splitext(file)
        if 'xlsx' not in ext.lower():
            continue
        print(file)
        wb = xl.load_workbook(f"./PROCESS/{file}")
        ws = wb[wb.sheetnames[0]]

        for row in range(row_start,ws.max_row+1):
            url = ''
            issn = ws.cell(row=row, column=issn_col).value.split(';')[0]

            c.execute("SELECT pdfversion FROM journals WHERE issn=:issn AND pdfversion='can'",{'issn':issn})
            tmp = c.fetchall()
            c.execute("SELECT pdfversion FROM journals WHERE essn=:essn AND pdfversion='can'",{'essn':issn})
            tmp += c.fetchall()

            if len(tmp) == 1:
                url = ws.cell(row=row, column=url_col).value
                if url == '':
                    continue
                try:
                    if url in [None,'']:
                        continue
                    driver.get(url)
                except:
                    print(f'FAIL: {url}')
                    continue
                try:
                    time.sleep(2)
                    driver.switch_to.alert().accept()
                    keyboard.press_and_release('enter')
                    time.sleep(10)
                    driver.refresh()
                except selenium.common.exceptions.NoAlertPresentException:
                    # NOT IMPLEMENTED
                    pass

                url = driver.current_url
                if 'can' == str(tmp[0][0]) and url != '':
                    try:
                        if 'spiedigitallibrary' in url:
                            # Subscription
                            continue
                        elif 'abt.ucpress' in url:
                            # Subscription
                            continue
                        elif 'earthquakespectra' in url:
                            # Subscription
                            continue
                        elif 'igi-global' in url:
                            # Login required.
                            continue
                        elif 'library.seg' in url:
                            # Purchase
                            continue
                        elif 'freshwater-science' in url:
                            continue
                        elif 'ingentaconnect' in url:
                            # Purchase
                            continue
                        elif 'mmnp-journal' in url:
                            # Purchase
                            continue
                        elif 'mining.archives' in url:
                            # Login
                            continue
                        elif 'jov.arvojournals' in url:
                            # Purchase
                            continue
                        # print(f'\t{row}:\n\t\t{url}\n\n')
                        if 'journals.uchicago' in url:
                            # Purchase/Request
                            url = uchicago(url,bs(driver.page_source,'html.parser'))
                            # continue
                        elif 'pubs.geoscienceworld' in url:
                            url = geoscienceworld(driver)
                        elif 'sciencedirect' in url:
                            url = sciencedirect(driver)
                        elif 'osapublishing' in url:
                            url = osa(url,bs(driver.page_source,'html.parser'))
                        elif 'iopscience.iop' in url:
                            # They've added protections against scrapping.
                            continue
                            url = iopscience(url,bs(driver.page_source,'html.parser'))
                        elif 'journals.aps' in url:
                            url = aps(driver.current_url,bs(driver.page_source,'html.parser'))
                            driver.get(url)
                            if 'Verification Required' in bs(driver.page_source,'html.parser').text:
                                input('PENDING...')
                        elif 'journals.plos' in url:
                            url = plos(url,bs(driver.page_source,'html.parser'))
                        elif 'www.jci.org' in url:
                            url = jci(driver)
                        elif 'hindawi' in url:
                            url = hindawi(url,bs(driver.page_source,'html.parser'))
                        elif 'pubs.rsc' in url:
                            url = pubs_rsc(url,bs(driver.page_source,'html.parser'))
                        elif 'nature.com' in url:
                            url = nature(url,bs(driver.page_source,'html.parser'))
                        elif 'bmccomplementalternmed' in url:
                            url = bmc(url,bs(driver.page_source,'html.parser'))
                        # elif 'advancesindifferenceequations.springeropen' in url:
                        #     url = springeropen(url,bs(driver.page_source,'html.parser'))
                        # elif 'link.springer':
                        #     url = link_springer(url,bs(driver.page_source,'html.parser'))
                        elif 'springer' in url:
                            try:
                                url = springer(url,bs(driver.page_source,'html.parser'))
                            except:
                                print(url)
                        elif 'academic.oup' in url:
                            try:
                                url = oxford(url,bs(driver.page_source,'html.parser'))
                            except IndexError:
                                continue
                        elif 'jnsbm' in url:
                            url = jnsbm(driver)
                        elif 'mdpi.com' in url:
                            url = mdpi(url,bs(driver.page_source,'html.parser'))
                        elif 'journals.sagepub' in url:
                            url = sage(url,bs(driver.page_source,'html.parser'))
                        elif 'projecteuclid' in url:
                            url = euclid(url,bs(driver.page_source,'html.parser'))
                            # driver.close()
                            # driver = webdriver.Chrome(chrome_options = options)
                        elif 'epubs.siam' in url:
                            url = siam(url,bs(driver.page_source,'html.parser'))
                        elif 'file.scirp' in url:
                            url = scirp(url,bs(driver.page_source,'html.parser'))
                        elif 'scripts.iucr' in url:
                            url = scripts_iucr(driver)
                        elif 'onlinelibrary.wiley' in url:
                            wiley(driver)
                        elif 'iumj.indiana' in url:
                            url = iumj(url,bs(driver.page_source,'html.parser'))
                        elif 'sciedupress' in url:
                            url = sciedupress(url,bs(driver.page_source,'html.parser'))
                        else:
                            print(url)
                            url = '' 
                    except IndexError:
                        print(f"ERROR: {url}")
                    except KeyboardInterrupt:
                        # print(f'FAIL: {url}')
                        # url = ''
                        print('\n\n\n\nINTERRUPTED... \n\n\n\n')
                        try:
                            driver.close()
                        except:
                            pass
                        driver = webdriver.Chrome(chrome_options = options)
                        continue

            elif len(tmp) > 1:
                print(f'MANUAL ENTRY: {issn}')
            if url != '':
                doi = ws.cell(row=row, column=doi_col).value.replace('/','~')
                for i in [':','\\','*','?','<','>','|','.']:
                    doi = doi.replace(i,'')
                # try:
                if len(os.listdir(f'{path}/Selenium_Download_TMP')) == 0:
                    driver.get(url)
                    time.sleep(2)
                if len(os.listdir(f'{path}/Selenium_Download_TMP')) == 0:
                    continue
                while '.pdf' not in [os.path.splitext(x)[1][-4:].lower() for x in os.listdir(f'{path}/Selenium_Download_TMP')]:
                    pass
                time.sleep(2)
                for i in os.listdir(f'{path}/Selenium_Download_TMP'):
                    _,ext = os.path.splitext(i)
                    if '.pdf' not in ext[-4:].lower():
                        while True:
                            try:
                                os.remove(i)
                                break
                            except PermissionError as e:
                                time.sleep(2)
                            except FileNotFoundError as e:
                                break
                download_name = os.listdir(f'{path}/Selenium_Download_TMP')[0]
                os.rename(f'{path}/Selenium_Download_TMP/{download_name}',f'{path}/{name}_{row}.pdf')
                for i in os.listdir(f'{path}/Selenium_Download_TMP'):
                    os.remove(f'{path}/Selenium_Download_TMP/{i}')
        row_start = 1

if __name__ == "__main__":
    main()